# coding:utf8

"""Usage:blankRowInserter.py <n> <m> <fileName>
n           Row number to start.
m           Number of blank rows.
fileName    Target Excel file.
"""

import os
import sys
import time
import openpyxl
import openpyxl.worksheet.worksheet as oww


def get_new_filename(filename):
    twice_tup = os.path.splitext(filename)
    return time.strftime('_treated%Y%m%d%H%M%S', time.localtime()).join(twice_tup)


if len(sys.argv) == 4:
    for i in (1, 2):
        if not sys.argv[i].isdigit() and sys.argv[i] != '0':
            print('Invalid integer number %s' % sys.argv[i])
            sys.exit(2)
    if not os.path.exists(sys.argv[3]):
        print('%s File not found' % sys.argv[3])
        sys.exit(2)
    if not sys.argv[3].endswith('.xlsx'):
        print('%s Invalid file extension' % sys.argv[3].split('.')[-1])
        sys.exit(2)

    wb: openpyxl.Workbook = openpyxl.load_workbook(sys.argv[3])
    try:
        ws_ori: oww.Worksheet = wb.active
        if hasattr(ws_ori, 'insert_rows') and hasattr(wb, 'copy_worksheet') and 0:
            ws_new = wb.copy_worksheet(ws_ori)
            ws_new.insert_rows(int(sys.argv[1]), int(sys.argv[2]))
        else:
            # copy only value
            ws_new = wb.create_sheet('%s Copy' % ws_ori.title, wb.index(ws_ori) + 1)
            range_obj_for_x = range(1, ws_ori.max_column + 1)
            row_offset = int(sys.argv[2])
            for y in range(1, int(sys.argv[1])):
                for x in range_obj_for_x:
                    ws_new.cell(y, x).value = ws_ori.cell(y, x).value
            for y in range(int(sys.argv[1]), ws_ori.max_row + 1):
                for x in range_obj_for_x:
                    ws_new.cell(y + row_offset, x).value = ws_ori.cell(y, x).value
        wb.save(get_new_filename(sys.argv[3]))
    finally:
        wb.close()
    print('Done')
else:
    print(__doc__)
    if len(sys.argv) == 2 and sys.argv[1] in ('-h', '--help'):
        sys.exit()
    else:
        sys.exit(2)
