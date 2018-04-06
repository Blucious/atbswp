# coding:utf8

import sys
from pprint import pprint as ppr_pprint
from openpyxl import load_workbook as pyxl_load_workbook

try:
    wb = pyxl_load_workbook('censuspopdata.xlsx')
except FileNotFoundError as err:
    print(err)
    sys.exit(2)
else:
    try:
        sheet = wb['Population by Census Tract']
    except KeyError as err:
        print(err)
        sys.exit(2)

    county_data = {}
    last_state = None
    last_county = None
    for row in sheet.iter_rows(min_col=2, min_row=2):
        state = row[0].value
        county = row[1].value
        pop = row[2].value

        state_data = county_data.setdefault(state, {})
        info_data = state_data.setdefault(county, {'pop': 0, 'tracts': 0})
        info_data['pop'] += int(pop)
        info_data['tracts'] += 1

    ppr_pprint(county_data)
    sys.exit()
finally:
    if hasattr(locals(), 'wb'):
        wb.close()
