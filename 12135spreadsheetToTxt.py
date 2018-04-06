# coding:utf8

"""Usage:spreadsheetToTxt.py <fileName>"""

import os
import sys
import openpyxl
from openpyxl.cell.cell import get_column_letter, column_index_from_string

BASE_DIR = os.path.curdir + os.path.sep + '12134testFile'


if len(sys.argv) > 1:
    try:
        wb = openpyxl.load_workbook(sys.argv[1])
    except Exception as err:
        print('[-]Error: %s' % err)
        sys.exit()

    try:
        ws = wb.active

        try:
            os.mkdir(BASE_DIR)
        except FileExistsError:
            pass

        for x in range(1, ws.max_column + 1):
            with open(os.path.join(BASE_DIR, '%s.txt' % get_column_letter(x)), 'w') as fio:
                for y in range(1, ws.max_row + 1):
                    v = ws.cell(y, x).value
                    fio.write('%s\n' % str('' if v is None else v))
    finally:
        wb.close()

    print('[+]Done')
else:
    print(__doc__)
