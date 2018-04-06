# coding:utf8

"""Usage:spreadsheetCellInverter.py <fileName>"""

import os
import sys
import time
import openpyxl


def get_new_filename(filename):
    twice_tup = os.path.splitext(filename)
    return time.strftime('_treated%Y%m%d%H%M%S', time.localtime()).join(twice_tup)


if len(sys.argv) == 2:
    try:
        wb = openpyxl.load_workbook(sys.argv[1])
    except Exception as err:
        print('[-]Error: %s' % err)
        sys.exit()

    ws_ori = wb.active
    ws_new = wb.create_sheet('%s Copy' % ws_ori.title, wb.index(ws_ori) + 1)
    sheet_data = []
    try:
        # Only value.
        for y in range(1, ws_ori.max_row + 1):
            sheet_data.append([])
            for x in range(1, ws_ori.max_column + 1):
                sheet_data[y - 1].append(ws_ori.cell(y, x).value)

        for y in range(1, ws_ori.max_row + 1):
            for x in range(1, ws_ori.max_column + 1):
                ws_new.cell(x, y).value = sheet_data[y - 1][x - 1]

        wb.save(get_new_filename(sys.argv[1]))
    finally:
        wb.close()
else:
    print(__doc__)
