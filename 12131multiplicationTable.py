# coding:utf8

"""Usage:multiplicationTable.py <n>"""

import sys
import openpyxl

FILE_NAME = 'multiplicationTable.xlsx'


if len(sys.argv) == 2 and sys.argv[1].isdigit():
    wb = openpyxl.Workbook()
    ws = wb.active
    n = int(sys.argv[1])
    for i in range(1, n + 1):
        ws.cell(row=1, column=i + 1).value = i
        ws.cell(row=i + 1, column=1).value = i
    for i in range(1, n + 1):
        for j in range(1, n + 1):
            ws.cell(row=i + 1, column=j + 1).value = i * j
    wb.save(FILE_NAME)
    wb.close()
else:
    print(__doc__)
