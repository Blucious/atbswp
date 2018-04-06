# coding:utf8
# updateProduce.py - Corrects costs in produce sales spreadsheet.

import openpyxl

# The produce types and their updated prices
PRICE_UPDATES = {
    'Celery': 1.19,
    'Garlic': 3.07,
    'Lemon': 1.27
}

print('Reading...')
wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']

print('Scanning...')
# Loop through the rows and update the prices.
for row_index in range(2, sheet.max_row + 1):  # skip the first row
    produce_name = sheet.cell(row=row_index, column=1).value
    if produce_name in PRICE_UPDATES:
        sheet.cell(row=row_index, column=2).value = PRICE_UPDATES[produce_name]
        print('[+]row %d' % row_index)

print('Saving...')
wb.save('updatedProduceSales.xlsx')
wb.close()
print('Done.')
