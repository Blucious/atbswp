# coding:utf8

"""Usage:excelToCsv.py"""

import os
import csv
import time
import openpyxl
import traceback

BASE_DIR = 'csv' + time.strftime('%Y%m%d%H%M%S')


os.makedirs(BASE_DIR, exist_ok=True)

try:
    for excel_fn in os.listdir(os.path.curdir):
        # Skip non-xlsx files, load the workbook object.
        if excel_fn.startswith('~$') or os.path.splitext(excel_fn)[1] != '.xlsx':
            continue

        wb = openpyxl.load_workbook(excel_fn)
        try:
            for ws_name in wb.sheetnames:
                # Loop through every sheet in the workbook.
                ws = wb[ws_name]

                # Create tbe CSV filename from the Excel filename and sheet title.
                csv_fn = '%s_%s.csv' % (os.path.splitext(excel_fn)[0], ws_name)
                
                # Create the csv.writer object for this CSV file.
                csv_file = open(os.path.join(BASE_DIR, csv_fn), 'w', newline='')
                try:
                    csv_writer = csv.writer(csv_file)

                    # Loop through every row in the sheet.
                    for row_num in range(1, ws.max_row + 1):
                        row_data = []  # append each cell to this list

                        # Loop through each cell in the row.
                        for col_num in range(1, ws.max_column + 1):
                            # Append each cell's data to rowData.
                            row_data.append(ws.cell(row_num, col_num).value)

                        # Write the rowData list to the CSV file.
                        csv_writer.writerow(row_data)
                finally:
                    csv_file.close()
                print('[+]%s <%s> -> %s done' % (excel_fn, ws_name, csv_fn))
        finally:
            wb.close()
        print('[+]File %s done' % excel_fn)
except BaseException as err:
    print('[-]Error: %s' % str(err))
    traceback.print_exc()
print('All done')
